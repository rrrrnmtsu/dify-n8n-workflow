"""
CROSS ROPPONGI Excel Parser API
n8n専用の軽量Excelパースサービス
"""

from flask import Flask, request, jsonify
from openpyxl import load_workbook
from io import BytesIO
import re
from datetime import datetime, timedelta

app = Flask(__name__)

def cell_value(sheet, cell_address):
    """単一セルの値を取得"""
    try:
        return sheet[cell_address].value
    except:
        return None

def range_value(sheet, range_address):
    """範囲セルの最初の非空白値を取得"""
    try:
        start_cell, end_cell = range_address.split(':')
        for row in sheet[range_address]:
            for cell in row:
                if cell.value is not None and cell.value != '':
                    return cell.value
        return None
    except:
        return None

def range_array(sheet, range_address):
    """範囲セルから配列を取得"""
    try:
        values = []
        for row in sheet[range_address]:
            for cell in row:
                if cell.value is not None and cell.value != '':
                    values.append(cell.value)
        return values
    except:
        return []

def extract_table_data(sheet, name_range, amount_range, person_range=None):
    """テーブル形式のデータを抽出（顧客名、金額、担当者）"""
    try:
        result = []

        # 範囲を行ごとに分割
        name_cells = list(sheet[name_range])
        amount_cells = list(sheet[amount_range])
        person_cells = list(sheet[person_range]) if person_range else None

        # 行数を取得
        rows = len(name_cells)

        for i in range(rows):
            # 顧客名を取得（複数列の可能性があるので最初の非空白値）
            name = None
            for cell in name_cells[i]:
                if cell.value is not None and cell.value != '':
                    name = str(cell.value)
                    break

            # 金額を取得
            amount = None
            for cell in amount_cells[i]:
                if cell.value is not None and cell.value != '':
                    amount = parse_currency(cell.value)
                    break

            # 担当者を取得（オプション）
            person = None
            if person_cells:
                for cell in person_cells[i]:
                    if cell.value is not None and cell.value != '':
                        person = str(cell.value)
                        break

            # 顧客名と金額が両方ある場合のみ追加
            if name and amount is not None:
                entry = {
                    "customer_name": name,
                    "amount": amount
                }
                if person:
                    entry["person_in_charge"] = person
                result.append(entry)

        return result
    except Exception as e:
        print(f"Error extracting table data: {e}")
        return []

def parse_currency(value):
    """通貨値を数値に変換"""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        cleaned = value.replace('¥', '').replace(',', '').strip()
        try:
            return float(cleaned)
        except:
            return None

    return None

def normalize_date(value, year, month, day):
    """日付を正規化"""
    if value is None:
        return f"{year}-{month}-{day}"

    # datetime型の場合
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    # 数値（Excelシリアル値）の場合
    if isinstance(value, (int, float)):
        excel_epoch = datetime(1899, 12, 30)
        date = excel_epoch + timedelta(days=value)
        return date.strftime('%Y-%m-%d')

    # 文字列の場合
    if isinstance(value, str):
        # 既にYYYY-MM-DD形式の場合
        if re.match(r'\d{4}-\d{2}-\d{2}', value):
            return value
        # その他の形式は試行
        try:
            parsed = datetime.strptime(value, '%Y/%m/%d')
            return parsed.strftime('%Y-%m-%d')
        except:
            pass

    # フォールバック: ファイル名から推測
    return f"{year}-{month}-{day}"

@app.route('/health', methods=['GET'])
def health():
    """ヘルスチェック"""
    return jsonify({"status": "ok", "service": "excel-parser"})

@app.route('/parse', methods=['POST'])
def parse_excel():
    """Excelファイルをパースして構造化データを返す"""

    try:
        # ファイル名とバイナリデータを取得
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']

        # クエリパラメータから元のファイル名を取得（n8n経由の場合）
        # なければアップロードされたファイル名を使用
        filename = request.args.get('filename', file.filename)

        # ファイル名から日付抽出
        date_match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
        if not date_match:
            return jsonify({"error": f"Cannot extract date from filename: {filename}"}), 400

        year, month, day = date_match.groups()
        target_sheet_name = str(int(day))  # "18"

        # Excelファイルを読み込み
        wb = load_workbook(BytesIO(file.read()), data_only=True)

        # 利用可能なシート名を確認
        available_sheets = wb.sheetnames

        if target_sheet_name not in available_sheets:
            return jsonify({
                "error": f"Sheet '{target_sheet_name}' not found",
                "available_sheets": available_sheets
            }), 404

        sheet = wb[target_sheet_name]

        # データ抽出
        extracted_data = {
            # メタ情報
            "source_file": filename,
            "sheet_name": target_sheet_name,
            "extracted_at": datetime.now().isoformat(),

            # 基本情報
            "business_date": normalize_date(range_value(sheet, 'P2:R2'), year, month, day),
            "total_customer_count": cell_value(sheet, 'O2'),
            "male_count": cell_value(sheet, 'T2'),
            "female_count": cell_value(sheet, 'U2'),

            # 売上サマリー
            "total_sales": parse_currency(range_value(sheet, 'K64:M65')),
            "cash_shortage": parse_currency(range_value(sheet, 'O64:O65')),

            # セクション別売上
            "section_sales": {
                "front": parse_currency(cell_value(sheet, 'F5')),
                "cloak_supplies": parse_currency(cell_value(sheet, 'F10')),
                "locker": parse_currency(cell_value(sheet, 'F14')),
                "bar1": parse_currency(cell_value(sheet, 'F15')),
                "bar2": parse_currency(cell_value(sheet, 'F16')),
                "bar3": parse_currency(cell_value(sheet, 'F17')),
                "bar4": parse_currency(cell_value(sheet, 'F18')),
                "vip1": parse_currency(cell_value(sheet, 'F32')),
                "vvip": parse_currency(cell_value(sheet, 'F33')),
                "party": parse_currency(cell_value(sheet, 'F48'))
            },

            # 未収金
            "receivables": {
                "uncollected": parse_currency(range_value(sheet, 'J61:K61')),
                "collected": parse_currency(range_value(sheet, 'Q61:R61'))
            },

            # VIP詳細
            "vip_details": {
                "vip_customers": [],
                "vvip_customers": []
            },

            # 決済別データ
            "payment_methods": {
                "front": {
                    "cash": parse_currency(cell_value(sheet, 'P5')),
                    "credit": parse_currency(cell_value(sheet, 'G5')),
                    "quicpay": parse_currency(cell_value(sheet, 'H5')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I5')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J5')),
                    "jppoint": parse_currency(cell_value(sheet, 'M5')),
                    "receivable": parse_currency(cell_value(sheet, 'N5'))
                },
                "cloak": {
                    "cash": parse_currency(cell_value(sheet, 'P10')),
                    "credit": parse_currency(cell_value(sheet, 'G10')),
                    "quicpay": parse_currency(cell_value(sheet, 'H10')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I10')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J10')),
                    "jppoint": parse_currency(cell_value(sheet, 'M10')),
                    "receivable": parse_currency(cell_value(sheet, 'N10'))
                },
                "locker": {
                    "cash": parse_currency(cell_value(sheet, 'P14')),
                    "credit": parse_currency(cell_value(sheet, 'G14')),
                    "quicpay": parse_currency(cell_value(sheet, 'H14')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I14')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J14')),
                    "jppoint": parse_currency(cell_value(sheet, 'M14')),
                    "receivable": parse_currency(cell_value(sheet, 'N14'))
                },
                "bar1": {
                    "cash": parse_currency(cell_value(sheet, 'P15')),
                    "credit": parse_currency(cell_value(sheet, 'G15')),
                    "quicpay": parse_currency(cell_value(sheet, 'H15')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I15')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J15')),
                    "jppoint": parse_currency(cell_value(sheet, 'M15')),
                    "receivable": parse_currency(cell_value(sheet, 'N15'))
                },
                "bar2": {
                    "cash": parse_currency(cell_value(sheet, 'P16')),
                    "credit": parse_currency(cell_value(sheet, 'G16')),
                    "quicpay": parse_currency(cell_value(sheet, 'H16')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I16')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J16')),
                    "jppoint": parse_currency(cell_value(sheet, 'M16')),
                    "receivable": parse_currency(cell_value(sheet, 'N16'))
                },
                "bar3": {
                    "cash": parse_currency(cell_value(sheet, 'P17')),
                    "credit": parse_currency(cell_value(sheet, 'G17')),
                    "quicpay": parse_currency(cell_value(sheet, 'H17')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I17')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J17')),
                    "jppoint": parse_currency(cell_value(sheet, 'M17')),
                    "receivable": parse_currency(cell_value(sheet, 'N17'))
                },
                "bar4": {
                    "cash": parse_currency(cell_value(sheet, 'P18')),
                    "credit": parse_currency(cell_value(sheet, 'G18')),
                    "quicpay": parse_currency(cell_value(sheet, 'H18')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I18')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J18')),
                    "jppoint": parse_currency(cell_value(sheet, 'M18')),
                    "receivable": parse_currency(cell_value(sheet, 'N18'))
                },
                "vip": {
                    "cash": parse_currency(cell_value(sheet, 'P32')),
                    "credit": parse_currency(cell_value(sheet, 'G32')),
                    "quicpay": parse_currency(cell_value(sheet, 'H32')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I32')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J32')),
                    "jppoint": parse_currency(cell_value(sheet, 'M32')),
                    "receivable": parse_currency(cell_value(sheet, 'N32'))
                },
                "vvip": {
                    "cash": parse_currency(cell_value(sheet, 'P33')),
                    "credit": parse_currency(cell_value(sheet, 'G33')),
                    "quicpay": parse_currency(cell_value(sheet, 'H33')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I33')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J33')),
                    "jppoint": parse_currency(cell_value(sheet, 'M33')),
                    "receivable": parse_currency(cell_value(sheet, 'N33'))
                },
                "party": {
                    "cash": parse_currency(cell_value(sheet, 'P48')),
                    "credit": parse_currency(cell_value(sheet, 'G48')),
                    "quicpay": parse_currency(cell_value(sheet, 'H48')),
                    "airpay_qr": parse_currency(cell_value(sheet, 'I48')),
                    "zentoshin": parse_currency(cell_value(sheet, 'J48')),
                    "jppoint": parse_currency(cell_value(sheet, 'M48')),
                    "receivable": parse_currency(cell_value(sheet, 'N48'))
                }
            },

            # VIPリスト
            "vip_list": [],

            # VVIPリスト
            "vvip_list": [],

            # 未収リスト
            "uncollected_list": [],

            # 未収回収リスト
            "collected_list": []
        }

        # VIP顧客データ（既存のvip_details用）
        vip_names = range_array(sheet, 'AA5:AA27')
        vip_amounts = range_array(sheet, 'AB5:AB27')

        for i in range(min(len(vip_names), len(vip_amounts))):
            if vip_names[i] and vip_amounts[i]:
                extracted_data['vip_details']['vip_customers'].append({
                    "name": str(vip_names[i]),
                    "amount": parse_currency(vip_amounts[i])
                })

        # VVIP顧客データ（既存のvip_details用）
        vvip_names = range_array(sheet, 'AA29:AA52')
        vvip_amounts = range_array(sheet, 'AB29:AB52')

        for i in range(min(len(vvip_names), len(vvip_amounts))):
            if vvip_names[i] and vvip_amounts[i]:
                extracted_data['vip_details']['vvip_customers'].append({
                    "name": str(vvip_names[i]),
                    "amount": parse_currency(vvip_amounts[i])
                })

        # 2. VIPリスト（新規シート用）
        extracted_data['vip_list'] = extract_table_data(sheet, 'AA5:AA27', 'AB5:AB27')

        # 3. VVIPリスト（新規シート用）
        extracted_data['vvip_list'] = extract_table_data(sheet, 'AA29:AA52', 'AB29:AB52')

        # 4. 未収リスト
        extracted_data['uncollected_list'] = extract_table_data(sheet, 'H54:I60', 'J54:J60', 'K54:K60')

        # 5. 未収回収リスト
        extracted_data['collected_list'] = extract_table_data(sheet, 'O54:P60', 'Q54:Q60', 'R54:R60')

        # データ検証
        validation = {
            "warnings": [],
            "errors": []
        }

        # 来客数の整合性チェック
        if extracted_data['male_count'] and extracted_data['female_count'] and extracted_data['total_customer_count']:
            expected_total = extracted_data['male_count'] + extracted_data['female_count']
            if expected_total != extracted_data['total_customer_count']:
                validation['warnings'].append(
                    f"来客数不一致: 男性({extracted_data['male_count']}) + "
                    f"女性({extracted_data['female_count']}) = {expected_total} "
                    f"≠ 総来客数({extracted_data['total_customer_count']})"
                )

        # 必須フィールドチェック
        if not extracted_data['business_date']:
            validation['errors'].append('営業日が取得できませんでした (P2:R2)')
        if not extracted_data['total_sales']:
            validation['warnings'].append('総売上が取得できませんでした (K64:M65)')

        extracted_data['validation'] = validation

        return jsonify(extracted_data)

    except Exception as e:
        return jsonify({
            "error": str(e),
            "type": type(e).__name__
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
